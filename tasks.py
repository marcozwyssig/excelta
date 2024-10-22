import os
import shutil
import pandas as pd
import diff_match_patch as dmp_module

from invoke import task
from openpyxl import Workbook, load_workbook
from jinja2 import Environment, FileSystemLoader

sourcepath = os.getcwd()

@task
def remove_existing_folder(c):
    """Remove the existing Sheets folder."""
    print("Removing existing Sheets folder...")
    path = os.path.join(sourcepath, 'Sheets')
    if os.path.exists(path):
        shutil.rmtree(path)

@task
def align_sheets(c):
    """Align sheet filenames by replacing spaces with underscores."""
    print("Aligning sheet filenames...")
    xlsx_files = [file for file in os.listdir(sourcepath) if file.endswith('xlsx')]
    for file in xlsx_files:
        new_file = file.replace(' ', '_')
        if new_file != file:
            os.rename(os.path.join(sourcepath, file), os.path.join(sourcepath, new_file))
            print(f'Renamed: {file} -> {new_file}')

@task(pre=[remove_existing_folder, align_sheets])
def export_sheets(c):
    """Export sheets by separating them into individual files."""
    print("Exporting sheets...")
    xlsx_files = [file for file in os.listdir(sourcepath) if file.endswith('xlsx')]
    for file in xlsx_files:
        wb = load_workbook(file)
        output_folder = os.path.join(sourcepath, 'Sheets', file[:-5])
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        for sheet_name in wb.sheetnames:
            sheet_file_name = os.path.join(output_folder, f"{sheet_name}.xlsx")
            ws = wb[sheet_name]
            new_wb = Workbook()
            new_ws = new_wb.create_sheet(title=sheet_name)

            headers = [cell.value for cell in ws[1]]
            for col_num, header in enumerate(headers, start=1):
                new_ws.cell(row=1, column=col_num, value=header)

            if 'Req ID' in headers:
                req_id_col = headers.index('Req ID') + 1
                req_version_col = len(headers) + 1
                new_ws.cell(row=1, column=req_version_col, value='Req Version')

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.column == req_id_col and cell.value:
                            if isinstance(cell.value, str) and '-' in cell.value:
                                parts = cell.value.split('-')
                                new_ws.cell(row=cell.row, column=req_id_col, value=parts[0])
                                new_ws.cell(row=cell.row, column=req_version_col, value=parts[1])
            else:
                for row in ws.iter_rows():
                    for cell in row:
                        new_ws[cell.coordinate] = cell.value

            if 'Sheet' in new_wb.sheetnames:
                new_wb.remove(new_wb['Sheet'])
            new_wb.save(sheet_file_name)

@task
def compare_sheets(c):
    """Compare exported sheets and generate an HTML report."""
    print("Comparing sheets...")
    output_folder = os.path.join(sourcepath, 'Sheets')
    data = [dir for dir in os.listdir(output_folder)]

    def extract_date(item):
        return item.split()[-1]

    data = sorted(data, key=extract_date)

    sheet_names = ["All"]

    pair_dict = {f"{data[i][-8:]}-{data[i + 1][-8:]}": (data[i], data[i + 1]) for i in range(len(data) - 1)}

    for name, pair in pair_dict.items():
        print(f"Comparing {name} - {pair}")
        for sheetname in sheet_names:
            filename = f"{sheetname}.xlsx"
            report_name = f"{sheetname}.html"
            try:
                compare_sheet(filename, name, report_name, name, pair, sourcepath)
            except Exception as e:
                print(f"An exception of type {type(e).__name__} occurred. Arguments: {e.args}")

def compare_sheet(filename, delta, report_name, name, pair, sourcepath):
    """Helper function to compare sheets."""
    def ensure_directory_exists(filepath):
        directory = os.path.dirname(filepath)
        if not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
        return filepath

    source_file = ensure_directory_exists(os.path.join(sourcepath, 'Sheets', pair[0], f"{filename}"))
    target_file = ensure_directory_exists(os.path.join(sourcepath, 'Sheets', pair[1], f"{filename}"))
    report_file = ensure_directory_exists(os.path.join(sourcepath, 'Changes', name, report_name))

    generate_html_report(delta, source_file, target_file, report_file)

def read_excel(file_path):
    """Read Excel file into a DataFrame."""
    return pd.read_excel(file_path, engine='openpyxl')

def generate_html_report(title, source_file, target_file, report_file):
    """Generate an HTML report comparing two Excel sheets using Jinja2."""
    
    # Load the DataFrames from the source and target Excel files
    source_df = pd.read_excel(source_file)
    target_df = pd.read_excel(target_file)
    
    # Compare the sheets
    comparison_columns, comparison_rows = compare_sheet_df(source_df, target_df)
    
    # Create a Jinja2 environment and specify the folder where templates are stored
    env = Environment(loader=FileSystemLoader('templates'))
    
    # Load the template
    template = env.get_template('report_template.html')
    
    # Render the template with comparison data
    rendered_html = template.render(title=title,columns=comparison_columns, rows=comparison_rows)
    
    # Write the rendered HTML to the report file
    with open(report_file, "w", encoding="utf-8") as file:
        file.write(rendered_html)

def compare_sheet_df(source_df, target_df):
    """Generate the HTML report showing the differences between two DataFrames, 
    with a new column indicating 'No Change' or 'Changed'."""
    
    dmp = dmp_module.diff_match_patch()
    
    def get_diff(source, target):
        diffs = dmp.diff_main(str(source), str(target))
        dmp.diff_cleanupSemantic(diffs)
        html_diff = dmp.diff_prettyHtml(diffs)
        return html_diff, diffs

    comparison_rows = []
    source_keys = source_df.iloc[:, 0]
    target_keys = target_df.iloc[:, 0]
    all_keys = pd.Series(list(set(source_keys).union(set(target_keys))))

    for key in all_keys:
        if key in source_keys.values:
            source_row = source_df[source_df.iloc[:, 0] == key].iloc[0]
        else:
            source_row = pd.Series([key] + [''] * (len(source_df.columns) - 1), index=source_df.columns)

        if key in target_keys.values:
            target_row = target_df[target_df.iloc[:, 0] == key].iloc[0]
        else:
            target_row = pd.Series([key] + [''] * (len(target_df.columns) - 1), index=target_df.columns)

        row_class = ""
        status = "None"  # Default to No Change
        if key not in target_keys.values:
            row_class = "table-danger"
            status = "CHANGED"  # Inserted row
        elif key not in source_keys.values:
            row_class = "table-success"
            status = "CHANGED"  # Deleted row
        else:
            row_class = ""

        combined_row = []
        row_changed = False  # Track if the row is changed
        for col in source_df.columns:
            if col == source_df.columns[0]:
                combined_row.append({'key': key, 'class': '', 'diff_html': key})
            else:
                diff_html, diffs = get_diff(source_row[col], target_row[col])
                cell_class = ""
                if any(d[0] != 0 for d in diffs):  # Detects any differences in the column
                    cell_class = "table-warning"
                    row_changed = True
                combined_row.append({'key': key, 'class': cell_class, 'diff_html': diff_html})

        if row_changed and status == "None":
            status = "CHANGED"  # Mark the row as changed if any column differs

        # Append the status to the row as the last column
        combined_row.append({'key': key, 'class': '', 'diff_html': status})
        comparison_rows.append({'row_class': row_class, 'combined_row': combined_row})

    # Add a new column header for the "No Change" or "Changed" status
    columns = list(source_df.columns) + ["Change Status"]
    return columns, comparison_rows
